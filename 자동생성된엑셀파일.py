import openpyxl
import random

# 샘플 제품 데이터 준비
product_names = [
    "스마트폰", "노트북", "태블릿", "스마트워치", "무선이어폰", 
    "블루투스 스피커", "공기청정기", "로봇청소기", "전기밥솥", "전자레인지"
]

# 새로운 워크북 생성
wb = openpyxl.Workbook()
ws = wb.active
ws.title = "전자제품 판매데이터"

# 헤더 추가
headers = ["제품ID", "제품명", "수량", "가격"]
for col, header in enumerate(headers, 1):
    ws.cell(row=1, column=col, value=header)

# 100개의 데이터 생성 및 추가
for row in range(2, 102):  # 2부터 101까지 (100개)
    product_id = f"P{str(row-1).zfill(3)}"  # P001, P002, ...
    product_name = random.choice(product_names)
    quantity = random.randint(1, 100)
    price = random.randint(50000, 2000000)  # 5만원 ~ 200만원
    
    ws.cell(row=row, column=1, value=product_id)
    ws.cell(row=row, column=2, value=product_name)
    ws.cell(row=row, column=3, value=quantity)
    ws.cell(row=row, column=4, value=price)

# 파일 저장
wb.save('products.xlsx')
print("데이터가 성공적으로 저장되었습니다.")